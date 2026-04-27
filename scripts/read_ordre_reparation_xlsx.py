"""
Lit la maquette ORDRE DE REPARATION (1).xlsx (openpyxl) et affiche
fusion de cellules, dimensions et toutes les cellules non vides.

Usage (à la racine du projet) :
  pip install openpyxl
  python scripts/read_ordre_reparation_xlsx.py
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Racine = parent du dossier scripts/
ROOT = Path(__file__).resolve().parents[1]
DEFAULT_FILE = ROOT / "ORDRE DE REPARATION (1).xlsx"


def main() -> int:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_FILE
    if not path.is_file():
        print(f"Fichier introuvable : {path}", file=sys.stderr)
        return 1

    wb = load_workbook(path, data_only=True)
    out: dict = {"path": str(path), "sheets": []}

    for name in wb.sheetnames:
        ws = wb[name]
        sheet_info: dict = {
            "name": name,
            "dimensions": f"{get_column_letter(1)}1:{get_column_letter(ws.max_column)}{ws.max_row}",
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "merged": [str(r) for r in ws.merged_cells.ranges],
            "column_widths": {
                get_column_letter(c): round(ws.column_dimensions[get_column_letter(c)].width or 0, 3)
                for c in range(1, min(ws.max_column + 1, 28))
                if (ws.column_dimensions[get_column_letter(c)].width or 0) > 0
            },
            "cells_non_vides": [],
        }
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=row, column=col).value
                if v is not None and str(v).strip() != "":
                    ref = f"{get_column_letter(col)}{row}"
                    sheet_info["cells_non_vides"].append(
                        {"ref": ref, "row": row, "col": col, "value": v}
                    )
        out["sheets"].append(sheet_info)

    print(json.dumps(out, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
